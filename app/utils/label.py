from collections.abc import Sequence
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from reportlab.graphics.barcode import code128
from reportlab.lib.pagesizes import mm
from reportlab.pdfgen import canvas

LABEL_WIDTH = 64 * mm
LABEL_HEIGHT = 15 * mm
PDF_LABEL_WIDTH = 50 * mm
PDF_LABEL_HEIGHT = 25 * mm
HALF_WIDTH = LABEL_WIDTH / 2
SPREADSHEET_FORMULA_PREFIXES = ("=", "+", "-", "@", "\t", "\r", "\n")


def _spreadsheet_literal(value: object) -> object:
    if isinstance(value, str) and value.startswith(SPREADSHEET_FORMULA_PREFIXES):
        return f"'{value}"
    return value


def _draw_single_label(c, item):
    """Draws ONE foldable label on the current PDF page"""

    left_margin = 1.8 * mm
    top = LABEL_HEIGHT - 2 * mm

    # =========================
    # FRONT (LEFT 32mm)
    # =========================
    c.setFont("Helvetica-Bold", 6)
    c.drawString(left_margin, top, item.name)

    c.setFont("Helvetica", 5.5)
    c.drawString(
        left_margin,
        top - 4 * mm,
        f"Purity: {item.purity}%",
    )

    # Display the fixed rate for unique items, otherwise display weight.
    weight_text = (
        f"Rate: {item.fixed_rate}" if item.category == "unique" else f"{item.net_weight} g"
    )
    c.drawString(
        left_margin,
        top - 8 * mm,
        weight_text if item.category == "unique" else f"Net: {weight_text}",
    )

    # =========================
    # BACK (RIGHT 32mm)
    # =========================
    if item.barcode:
        back_x = HALF_WIDTH + 1.5 * mm

        # Barcode number
        c.setFont("Helvetica", 5)
        c.drawString(
            back_x,
            LABEL_HEIGHT - 3 * mm,
            item.barcode,
        )

        # Barcode
        barcode = code128.Code128(
            item.barcode,
            barHeight=7 * mm,
            barWidth=0.38,
            humanReadable=False,
        )

        barcode.drawOn(
            c,
            back_x,
            1.5 * mm,
        )


def generate_single_label_pdf(item) -> bytes:
    """Returns PDF bytes for ONE label"""
    buffer = BytesIO()
    c = canvas.Canvas(buffer, pagesize=(LABEL_WIDTH, LABEL_HEIGHT))

    _draw_single_label(c, item)

    c.showPage()
    c.save()

    buffer.seek(0)
    return buffer.read()


def _draw_pdf_batch_label(c, item):
    center_x = PDF_LABEL_WIDTH / 2
    first_line_y = PDF_LABEL_HEIGHT - 6 * mm
    second_line_y = first_line_y - 5 * mm
    barcode_y = 4 * mm

    c.setFont("Helvetica-Bold", 8)
    c.drawCentredString(center_x, first_line_y, item.name or "")

    c.setFont("Helvetica", 6.5)
    if getattr(item, "category", None) == "unique":
        weight_text = "Fixed price"
    else:
        net_weight = getattr(item, "net_weight", None)
        weight_text = f"{net_weight} g" if net_weight is not None else "-"

    purity_val = getattr(item, "purity", None)
    purity_text = f"{purity_val}%" if purity_val not in (None, 0.0, 0, 0.00) else ""
    if getattr(item, "category", None) == "unique":
        mc_text = f"Rate: {getattr(item, 'fixed_rate', 0)}"
    else:
        making_charge = getattr(item, "making_charge", None)
        mc_text = f"MC: {making_charge}" if making_charge is not None else "MC: -"

    second_line_parts = [weight_text, mc_text]
    if purity_text:
        second_line_parts.append(purity_text)

    c.drawCentredString(
        center_x,
        second_line_y,
        " • ".join(part for part in second_line_parts if part),
    )

    if getattr(item, "barcode", None):
        barcode_width = 1.0
        barcode = code128.Code128(
            item.barcode,
            barHeight=5 * mm,
            barWidth=barcode_width,
            humanReadable=False,
        )
        barcode_x = (PDF_LABEL_WIDTH - barcode.width) / 2
        barcode.drawOn(c, barcode_x, barcode_y)

        c.setFont("Helvetica", 6)
        c.drawCentredString(
            center_x,
            barcode_y - 3 * mm,
            item.barcode,
        )


def generate_batch_labels_pdf(items: Sequence) -> bytes:
    """
    Generates a PDF where each page contains one label with:
    - name on the first line
    - net weight, making charge, and purity on the second line
    - barcode graphic on the third line with barcode text below
    """

    PAGE_WIDTH = PDF_LABEL_WIDTH  # 50mm
    PAGE_HEIGHT = PDF_LABEL_HEIGHT  # 25mm

    buffer = BytesIO()
    c = canvas.Canvas(buffer, pagesize=(PAGE_WIDTH, PAGE_HEIGHT))

    for item in items:
        _draw_pdf_batch_label(c, item)
        c.showPage()

    c.save()
    buffer.seek(0)
    return buffer.read()


def generate_batch_labels_xlsx(items: Sequence) -> bytes:
    """
    Generates an XLSX file where:
    - Each row contains 3 items (Name, Purity, Charge or Rate, Weight, Barcode)
    - If items are not in multiples of 3, dummy1/dummy2 are added
    - Columns: Name 1, Name 2, Name 3, Purity 1, Purity 2, Purity 3,
      Charge 1, Charge 2, Charge 3, Wt. 1, Wt. 2, Wt. 3, Barcode 1, Barcode 2,
      Barcode 3
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Labels"

    # Set column widths
    for col in range(1, 16):
        ws.column_dimensions[chr(64 + col)].width = 15

    # Add header row
    headers = [
        "Name 1",
        "Name 2",
        "Name 3",
        "Purity 1",
        "Purity 2",
        "Purity 3",
        "Charge 1",
        "Charge 2",
        "Charge 3",
        "Wt. 1",
        "Wt. 2",
        "Wt. 3",
        "Barcode 1",
        "Barcode 2",
        "Barcode 3",
    ]
    ws.append(headers)

    # Style header
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    items_list = list(items)

    # Add data rows (3 items per row)
    for i in range(0, len(items_list), 3):
        row_data = []

        # Names (columns 1-3)
        for j in range(3):
            if i + j < len(items_list):
                row_data.append(_spreadsheet_literal(items_list[i + j].name))
            else:
                row_data.append("")

        # Purities (columns 4-6)
        for j in range(3):
            if i + j < len(items_list):
                item = items_list[i + j]
                purity_text = (
                    f"{item.purity}%" if getattr(item, "purity", None) in (92.5, 99.9) else ""
                )
                row_data.append(purity_text)
            else:
                row_data.append("")

        # Making charge or unique fixed rate (columns 7-9)
        for j in range(3):
            if i + j < len(items_list):
                item = items_list[i + j]
                value = (
                    getattr(item, "fixed_rate", 0)
                    if getattr(item, "category", None) == "unique"
                    else getattr(item, "making_charge", None)
                )
                mc_text = f"{value}" if value is not None else ""
                row_data.append(mc_text)
            else:
                row_data.append("")

        # Weights (columns 10-12)
        for j in range(3):
            if i + j < len(items_list):
                item = items_list[i + j]
                if hasattr(item, "category") and item.category == "unique":
                    weight_text = "unique"
                elif hasattr(item, "net_weight") and item.net_weight:
                    weight_text = f"{item.net_weight} g"
                else:
                    weight_text = ""
                row_data.append(weight_text)
            else:
                row_data.append("")

        # Barcodes (columns 13-15)
        for j in range(3):
            if i + j < len(items_list):
                row_data.append(_spreadsheet_literal(items_list[i + j].barcode or ""))
            else:
                row_data.append("")

        ws.append(row_data)

    # Center align all data cells
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=15):
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center")

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()
